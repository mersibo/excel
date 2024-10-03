from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import ContentType, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, FSInputFile
import pandas as pd
from datetime import datetime, timedelta
import os
import psycopg2
from openpyxl import Workbook, load_workbook
from copy import copy
from openpyxl.worksheet.filters import AutoFilter

bot = Bot('7317608175:AAEozzgOB10iAZKuHtVxd6YDik0LdhOn-PA')
dp = Dispatcher()

TEMP_FOLDER = "uploads"

def save_to_db(data):
    """
    Сохраняет данные в базу данных PostgreSQL.
    """
    conn = psycopg2.connect(
        dbname='timesheets',
        user='postgres',
        password='admin',
        host='localhost',
        port='5432'
    )
    cursor = conn.cursor()

    insert_query = """
    INSERT INTO timesheets (date, name, role, incident, task, task_type, task_number, hours, direction)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    
    for row in data.itertuples(index=False):
        cursor.execute(insert_query, row)
    
    conn.commit()
    cursor.close()
    conn.close()

# Убедимся, что папка для временных файлов существует
if not os.path.exists(TEMP_FOLDER):
    os.makedirs(TEMP_FOLDER)

def merge_timesheets_with_styles(week_files, output_path):
    # Загружаем первый файл и используем его как основу
    template_wb = load_workbook(week_files[0])
    merged_ws = template_wb.active  # используем активный лист в первом файле

    # Копируем строки с данными и стилями из всех остальных файлов за неделю
    for file in week_files[1:]:
        temp_wb = load_workbook(file)
        
        # Проверяем наличие листа с именем 'Таймшиты'
        if 'Таймшиты' in temp_wb.sheetnames:
            temp_ws = temp_wb['Таймшиты']
            
            # Копируем все строки, кроме заголовочной строки
            copy_rows_with_styles(temp_ws, merged_ws, start_row=3, end_row=temp_ws.max_row, start_col=1, end_col=temp_ws.max_column)
        else:
            print(f"Warning: Worksheet 'Таймшиты' does not exist in file {file}")

    # Применяем фильтры и сохраняем итоговый файл
    merged_ws.auto_filter.ref = merged_ws.auto_filter.ref or "A2:I2"  # Применяем фильтр к строке заголовков
    template_wb.save(output_path)


def copy_rows_with_styles(src_ws, dest_ws, start_row=1, end_row=None, start_col=1, end_col=None):
    """
    Копируем строки и их стили из одного листа на другой.
    """
    if end_row is None:
        end_row = src_ws.max_row
    if end_col is None:
        end_col = src_ws.max_column
    
    for row in src_ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        dest_row = dest_ws.max_row + 1  # Следующая строка для записи в целевой лист
        for col_index, cell in enumerate(row, start=start_col):
            new_cell = dest_ws.cell(row=dest_row, column=col_index)
            new_cell.value = cell.value

            # Копируем стиль
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)


def apply_template_styles_and_filters(template_ws, dest_ws):
    """
    Применяем стили и фильтры из шаблона на лист назначения.
    """
    # Применение автофильтра
    dest_ws.auto_filter.ref = template_ws.auto_filter.ref

    # Применение стилей для заголовков
    for col_index, template_cell in enumerate(template_ws[2], start=1):
        dest_cell = dest_ws.cell(row=2, column=col_index)
        if template_cell.has_style:
            dest_cell.font = copy(template_cell.font)
            dest_cell.border = copy(template_cell.border)
            dest_cell.fill = copy(template_cell.fill)
            dest_cell.number_format = copy(template_cell.number_format)
            dest_cell.protection = copy(template_cell.protection)
            dest_cell.alignment = copy(template_cell.alignment)


@dp.message(Command('start'))
async def start(message: types.Message):
    main_keyboard = ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="Получить таймшиты")],
        [KeyboardButton(text="Получить сводные данные")]
    ], resize_keyboard=True)
    await message.answer('Отправьте файл в формате .xlsx или выберите действие:', reply_markup=main_keyboard)


@dp.message(F.content_type == ContentType.DOCUMENT)
async def xlsx_file(message: types.Message):
    file = message.document
    if not file.file_name.endswith('.xlsx'):
        await message.reply("Файл должен быть в формате .xlsx")
        return

    file_info = await bot.get_file(file.file_id)
    file_path = file_info.file_path

    # Определение текущей недели
    current_week = datetime.now().isocalendar()[1]
    current_year = datetime.now().year

    # Создание папки для файлов текущей недели
    week_folder = os.path.join(TEMP_FOLDER, f"{current_year}_week_{current_week}")
    if not os.path.exists(week_folder):
        os.makedirs(week_folder)

    local_file_path = os.path.join(week_folder, file.file_name)
    await bot.download_file(file_path, local_file_path)
    await message.reply(f"Файл сохранен как {file.file_name}")

    df = pd.read_excel(local_file_path, header=1)

    required_columns = ['Дата', 'Фамилия Имя', 'Роль', 'Инцидент', 'Задача', 'Вид задачи', '№ задачи в Битриксе', 'Кол-во часов', 'Направление']

    if list(df.columns) != required_columns:
        await message.reply('Ошибка: Неверные атрибуты')
        os.remove(local_file_path)
        return

    errors = []
    for index, row in df.iterrows():
        if row.isnull().any():
            errors.append(f"Ошибка: Пустые значения в строке {index + 3}")

        try:
            date = pd.to_datetime(row['Дата'], format='%d.%m.%Y')
        except ValueError:
            errors.append(f"Ошибка: Неверный формат даты в строке {index + 3}")

        if pd.to_datetime(row['Дата']).isocalendar()[1] != current_week:
            errors.append(f"Ошибка: Дата в строке {index + 3} не попадает в текущую неделю")

    if errors:
        await message.reply("\n".join(errors))
        os.remove(local_file_path)
        return

    total_hours = df['Кол-во часов'].sum()

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="Да", callback_data='confirm'),
            InlineKeyboardButton(text="Нет", callback_data='cancel')
        ]
    ])

    await message.reply(f"Общее количество часов за неделю: {total_hours}\nПодтвердить загрузку данных?", reply_markup=keyboard)


@dp.callback_query(F.data.in_(['confirm', 'cancel']))
async def handle_confirmation(callback_query: CallbackQuery):
    if callback_query.data == 'confirm':
        current_week = datetime.now().isocalendar()[1]
        current_year = datetime.now().year
        week_folder = os.path.join(TEMP_FOLDER, f"{current_year}_week_{current_week}")
        local_file_path = os.path.join(week_folder, 'file.xlsx')  # используйте правильный путь к файлу

        df = pd.read_excel(local_file_path, header=1)
        
        df = df[1:]

        save_to_db(df)
        await callback_query.message.answer("Данные сохранены и записаны в базу данных")
    else:
        await callback_query.message.answer("Загрузка данных отменена")

    await callback_query.answer()


@dp.message(F.text == "Получить таймшиты")
async def get_timesheets(message: types.Message):
    current_week = datetime.now().isocalendar()[1]
    current_year = datetime.now().year
    week_folder = os.path.join(TEMP_FOLDER, f"{current_year}_week_{current_week}")

    if not os.path.exists(week_folder):
        await message.reply("Нет загруженных файлов за текущую неделю.")
        return

    # Фильтруем файлы, чтобы выбрать только те, которые содержат в названии "Таймшиты"
    week_files = [os.path.join(week_folder, f) for f in os.listdir(week_folder) if "Таймшиты" in f]

    if not week_files:
        await message.reply("Нет файлов с таймшитами для объединения.")
        return

    combined_file_path = os.path.join(week_folder, f"combined_timesheets_week_{current_week}.xlsx")

    # Объединение файлов
    merge_timesheets_with_styles(week_files, combined_file_path)

    await message.reply_document(FSInputFile(combined_file_path))


@dp.message(F.text == "Получить сводные данные")
async def get_summary(message: types.Message):
    current_week = datetime.now().isocalendar()[1]
    current_year = datetime.now().year
    week_folder = os.path.join(TEMP_FOLDER, f"{current_year}_week_{current_week}")

    combined_file_path = os.path.join(week_folder, f"combined_timesheets_week_{current_week}.xlsx")

    if not os.path.exists(combined_file_path):
        await message.reply("Нет сводного файла за текущую неделю.")
        return

    # Читаем сводный файл
    combined_df = pd.read_excel(combined_file_path, header=1)

    # Группировка данных по "Фамилия Имя" и "Инцидент" с суммированием количества часов
    summary_df = combined_df.groupby(['Фамилия Имя', 'Инцидент'])['Кол-во часов'].sum().reset_index()

    # Переименовываем столбцы
    summary_df.columns = ['Фамилия Имя', 'Инцидент', 'Кол-во часов']

    # Рассчитываем общую сумму часов
    total_hours = summary_df['Кол-во часов'].sum()

    # Добавляем строку с общей суммой в конец DataFrame
    total_row = pd.DataFrame([{'Фамилия Имя': 'ИТОГО', 'Инцидент': '', 'Кол-во часов': total_hours}])
    summary_df = pd.concat([summary_df, total_row], ignore_index=True)

    # Сохраняем итоговый файл с общей суммой
    summary_file_path = os.path.join(week_folder, f"summary_week_{current_week}.xlsx")
    summary_df.to_excel(summary_file_path, index=False)

    # Рассчитываем сумму часов по инцидентам
    incident_hours = summary_df.groupby('Инцидент')['Кол-во часов'].sum().reset_index()

    # Формируем сообщение с результатами
    result_message = f"Общая сумма часов: {total_hours}\n\nЧасы по инцидентам:\n"
    for index, row in incident_hours.iterrows():
        if row['Инцидент']:  # Пропускаем строку с общей суммой
            result_message += f"{row['Инцидент']}: {row['Кол-во часов']} часов\n"

    await message.reply(result_message)
    await message.reply_document(FSInputFile(summary_file_path))

if __name__ == '__main__':
    dp.run_polling(bot)